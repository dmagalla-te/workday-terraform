import openpyxl
from pydantic import ValidationError
from typing import List, Union
from models.a2s_model import AgentToServerTestCreate
from models.http_model import HttpServerTestCreate
from models.pageload_model import PageLoadTestCreateTr
from models.dns_model import DNSTestCreate
from models.bgp_model import BGPTestCreate

# Assuming you already have the Pydantic models (HttpServerTest, PageLoadTest, AgentToServerTest)

def read_excel_with_sheets(file_path: str): #-> List[Union[HttpServerTestCreate, PageLoadTest, AgentToServerTestCreate]]
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    #print(workbook.sheetnames)

    # A list to hold the created test objects
    test_objects = []

    # Sheet to model mapping
    sheet_model_mapping = {
        "http-server": HttpServerTestCreate,
        "page-load": PageLoadTestCreateTr,
        "agent-to-server": AgentToServerTestCreate,
        "dnssec": DNSTestCreate,
        "dns-trace": DNSTestCreate,
        "bgp":BGPTestCreate
    }

    # Iterate through each sheet
    for sheet_name, model in sheet_model_mapping.items():
        
        if sheet_name not in workbook.sheetnames:
        
            print(f"Sheet '{sheet_name}' not found in the workbook")
            continue

        # Select the sheet
        sheet = workbook[sheet_name]

        # Get the headers (assuming the first row contains headers)
        #headers = [cell.value for cell in sheet[6]]
        headers = [cell.value for cell in sheet[6][:18]]

        # Iterate over the rows (skip the header row)
        for row in sheet.iter_rows(min_row=8, max_col=sheet.max_column,values_only=True):


            if row[1] == None: # si la row ya no tiene info entonces dejamos de iterar en esa sheet
                break
            
            # If it has information -> Convert the row to a dictionary (matching headers with row data)
            row_data = dict(zip(headers, row))
            row_data = {k: v for k, v in row_data.items() if k is not None}


            # Try to create an instance of the Pydantic model
            try:
                test_object = model(**row_data)
                test_objects.append(test_object)

            except ValidationError as e:
                
                print(f"Validation error in sheet '{sheet_name}', row {row_data}: {e}")

    # vamos a sacar en que account groups podemos hacer bulk:
    accounts = set()
    sheet = workbook["http-server"]
    # Iterate over the AZ column starting at row 8
    for row in range(8, sheet.max_row + 1):  # sheet.max_row gives the last row with data
        
        cell_value = sheet[f'AZ{row}'].value  # Get the value of each cell in column AZ
        
        if cell_value:  # If there's data in the cell
        
            label_name, acc_name = cell_value.split("-->")
            acc_name = acc_name.strip()
            accounts.add(acc_name)

            
    return test_objects, accounts



